import openpyexcel
from openpyxl.utils import get_column_letter, column_index_from_string

##Practise scripts for working withXL : Simon Nov 2020

#open workbook as object
wb = openpyexcel.load_workbook('c:\\python projects\\boring_files\\example.xlsx')

type(wb)

sheet = wb['Sheet1']

#check python object class
print(type(sheet))

print(sheet)
print(sheet.title)


#sheet['A1']
#sheet['b2']

#get cell value from worksheet
print(sheet['A1'].value)
print(sheet['b2'].value)

c = sheet['B1']

print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)

anotherSheet = wb.active

print(anotherSheet)

#getting cells from sheets
# Get the row, column, and value from the cell.
#
row1 = 'Row %s, Column %s is %s' % (c.row, c.column, c.value)
print(row1)

print(sheet['C1'].value)

####
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

#looping over cells
for i in range(1,8,2):
    print(i,sheet.cell(row=i, column=2).value)


print('max row is: ', sheet.max_row)
print('max column is: ', sheet.max_column)

#convert between column letter and number
print('convert column number to letter')
print(get_column_letter(1))

##get using max_column method to get letter
print (get_column_letter(sheet.max_column))
s = get_column_letter(sheet.max_column)

print(s)
print(column_index_from_string(s))


##getting rows and columns from sheets

print(tuple(sheet['A1':'C3']) )

print('*' * 50)
print('\n' * 3)

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
